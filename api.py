import os
import tempfile
import traceback
from datetime import datetime
from typing import List
from fastapi import FastAPI, File, UploadFile, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from xml_extractor import CFDIXMLExtractor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart3D, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.colors import ColorChoice
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.series import DataPoint

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def cleanup(path: str):
    if os.path.exists(path):
        os.remove(path)

@app.get("/")
async def health_check():
    return {"status": "online"}

@app.post("/api/analizar")
async def analizar_facturas(background_tasks: BackgroundTasks, files: List[UploadFile] = File(...)):
    try:
        resultados = []
        total_facturas = 0
        total_riesgo_monetario = 0.0
        total_errores = 0
        count_ok = 0
        count_riesgo = 0
        count_error = 0

        for file in files:
            if not file.filename.lower().endswith('.xml'):
                continue
            
            total_facturas += 1
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xml") as tmp_xml:
                tmp_xml.write(await file.read())
                tmp_xml_path = tmp_xml.name

            try:
                extractor = CFDIXMLExtractor(tmp_xml_path)
                rfc, uuid, subtotal, iva_d, isr_d = extractor.extract_data()
                iva_e, isr_e = extractor.validate_taxes(subtotal)
                
                dif_iva = abs(iva_e - iva_d)
                dif_isr = abs(isr_e - isr_d)
                discrepancia_total = dif_iva + dif_isr

                dictamen = "Sin discrepancias"
                if discrepancia_total > 0.10:
                    dictamen = "RIESGO FISCAL DETECTADO"
                    total_riesgo_monetario += discrepancia_total
                    count_riesgo += 1
                elif isinstance(rfc, str) and "Error" in rfc:
                    dictamen = "ERROR ESTRUCTURAL"
                    total_errores += 1
                    count_error += 1
                elif subtotal == 0.0:
                    dictamen = "ANOMALÍA: Sin subtotal"
                    total_errores += 1
                    count_error += 1
                else:
                    count_ok += 1

                resultados.append([
                    file.filename, uuid, rfc, subtotal, iva_d, iva_e, isr_d, isr_e, dictamen
                ])
            finally:
                if os.path.exists(tmp_xml_path):
                    os.remove(tmp_xml_path)

        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard de Auditoría"
        ws.sheet_view.showGridLines = False

        fill_blue = PatternFill(start_color="0F243E", end_color="0F243E", fill_type="solid")
        font_white_bold = Font(color="FFFFFF", bold=True, size=11)
        font_title = Font(color="FFFFFF", bold=True, size=16)
        align_center_no_wrap = Alignment(horizontal="center", vertical="center", wrap_text=False)
        thin_border = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
                             top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))

        ws.merge_cells('B2:J2')
        title_cell = ws['B2']
        title_cell.value = "DICTAMEN EJECUTIVO DE AUDITORÍA PREVENTIVA (ART. 30-B CFF)"
        title_cell.fill = fill_blue
        title_cell.font = font_title
        title_cell.alignment = align_center_no_wrap
        ws.row_dimensions[2].height = 40

        metricas = [
            ("B4", "B5", "FACTURAS PROCESADAS", total_facturas),
            ("E4", "E5", "RIESGO FISCAL TOTAL", total_riesgo_monetario),
            ("H4", "H5", "ERRORES ESTRUCTURALES", total_errores)
        ]
        for h_cell, v_cell, text, val in metricas:
            for c in [h_cell, v_cell]:
                ws[c].fill = fill_blue
                ws[c].font = font_white_bold
                ws[c].alignment = align_center_no_wrap
            ws[h_cell] = text
            ws[v_cell] = val
            if "RIESGO" in text:
                ws[v_cell].number_format = '"$"#,##0.00_-'
        
        ws.row_dimensions[4].height = 25
        ws.row_dimensions[5].height = 25

        headers = [
            "Archivo", "UUID (Folio Fiscal)", "RFC Emisor", "Subtotal Base", 
            "IVA Declarado", "IVA Esperado", "ISR Declarado", "ISR Esperado", "Dictamen Legal"
        ]
        for col_idx, text in enumerate(headers, start=2):
            cell = ws.cell(row=8, column=col_idx, value=text)
            cell.fill = fill_blue
            cell.font = font_white_bold
            cell.alignment = align_center_no_wrap
            cell.border = thin_border
        ws.row_dimensions[8].height = 30

        for row_idx, data in enumerate(resultados, start=9):
            ws.row_dimensions[row_idx].height = 30
            for col_idx, value in enumerate(data, start=2):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = align_center_no_wrap
                if 5 <= col_idx <= 9:
                    cell.number_format = '"$"#,##0.00_-'
                
                if col_idx == 10:
                    if value == "Sin discrepancias":
                        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                        cell.font = Font(color="274E13", bold=True)
                    elif "RIESGO" in value:
                        cell.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
                        cell.font = Font(color="990000", bold=True)
                    else: 
                        cell.fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
                        cell.font = Font(color="B45F06", bold=True)

        ws.auto_filter.ref = f"B8:J{8 + len(resultados)}"

        ws_data = wb.create_sheet("DatosGrafico")
        # Punto 1: Hoja oculta
        ws_data.sheet_state = 'hidden'
        g_data = [["Estado", "Cant"], ["Sin discrepancias", count_ok], ["Riesgo Fiscal", count_riesgo], ["Error / Anomalía", count_error]]
        for r in g_data: ws_data.append(r)

        chart = PieChart3D()
        chart.add_data(Reference(ws_data, min_col=2, min_row=1, max_row=4), titles_from_data=True)
        chart.set_categories(Reference(ws_data, min_col=1, min_row=2, max_row=4))
        chart.title = "Distribución de Resultados Fiscales"
        chart.legend.position = 'b'
        
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = False
        chart.dataLabels.showCatName = False
        chart.dataLabels.showLeaderLines = True
        
        cp = CharacterProperties(solidFill=ColorChoice(srgbClr='FFFFFF'), b=True, sz=1100)
        chart.dataLabels.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
        
        # Colores de la gráfica
        s = chart.series[0]
        pt0 = DataPoint(idx=0)
        pt0.graphicalProperties = GraphicalProperties(solidFill="274E13") # Verde
        pt1 = DataPoint(idx=1)
        pt1.graphicalProperties = GraphicalProperties(solidFill="990000") # Rojo
        pt2 = DataPoint(idx=2)
        pt2.graphicalProperties = GraphicalProperties(solidFill="B45F06") # Naranja/Amarillo
        s.dPt = [pt0, pt1, pt2]
        
        ws.add_chart(chart, "L4")

        # Ajuste de anchos para métricas superiores (B, E, H)
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['H'].width = 35
        
        # Ajuste dinámico para el resto de columnas de la tabla
        for col in range(2, 11):
            column_letter = get_column_letter(col)
            if column_letter not in ['B', 'E', 'H']:
                max_length = 0
                for row in ws.iter_rows(min_row=8, max_row=8 + len(resultados), min_col=col, max_col=col):
                    for cell in row:
                        if cell.value:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                ws.column_dimensions[column_letter].width = max_length + 5

        ws.column_dimensions['A'].width = 3

        timestamp = datetime.now().strftime("%d-%m-%Y_%H%Mhrs")
        final_name = f"Dictamen_Ejecutivo_Art30B_{timestamp}.xlsx"

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_xlsx:
            wb.save(tmp_xlsx.name)
            output_path = tmp_xlsx.name

        background_tasks.add_task(cleanup, output_path)
        return FileResponse(output_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=final_name)

    except Exception as e:
        print("ERROR EN API:")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.environ.get("PORT", 8000)))
